"""Fetch actual MLB results + historical odds for a date range and score model predictions.

Usage:
    python -m pipeline.backtest_week --start 2026-05-08 --end 2026-05-14
"""
from __future__ import annotations

import argparse
import logging
from datetime import date, timedelta
from pathlib import Path

import numpy as np
import pandas as pd
import requests
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.chart import BarChart, Reference

import config
from ingestion.odds_api import fetch_historical_dates
from edge.vig_removal import american_to_prob

logger = logging.getLogger(__name__)

_MLB_SCHEDULE = "https://statsapi.mlb.com/api/v1/schedule"

_MLB_TO_RETRO: dict[str, str] = {
    "ATH": "OAK", "AZ": "ARI",  "CHC": "CHN", "CWS": "CHA",
    "KC":  "KCA", "LAA": "ANA", "LAD": "LAN", "NYM": "NYN",
    "NYY": "NYA", "SD":  "SDN", "SF":  "SFN", "STL": "SLN",
    "TB":  "TBA", "WSH": "WAS",
}

_FULL_TO_RETRO: dict[str, str] = {
    "Arizona Diamondbacks": "ARI", "Atlanta Braves": "ATL",
    "Baltimore Orioles": "BAL", "Boston Red Sox": "BOS",
    "Chicago Cubs": "CHN", "Chicago White Sox": "CHA",
    "Cincinnati Reds": "CIN", "Cleveland Guardians": "CLE",
    "Colorado Rockies": "COL", "Detroit Tigers": "DET",
    "Houston Astros": "HOU", "Kansas City Royals": "KCA",
    "Los Angeles Angels": "ANA", "Los Angeles Dodgers": "LAN",
    "Miami Marlins": "MIA", "Milwaukee Brewers": "MIL",
    "Minnesota Twins": "MIN", "New York Mets": "NYN",
    "New York Yankees": "NYA", "Oakland Athletics": "OAK",
    "Philadelphia Phillies": "PHI", "Pittsburgh Pirates": "PIT",
    "San Diego Padres": "SDN", "San Francisco Giants": "SFN",
    "Seattle Mariners": "SEA", "St. Louis Cardinals": "SLN",
    "Tampa Bay Rays": "TBA", "Texas Rangers": "TEX",
    "Toronto Blue Jays": "TOR", "Washington Nationals": "WAS",
    "Athletics": "OAK",
}

CONFIDENCE_THRESHOLDS = [0.50, 0.55, 0.60, 0.65, 0.70]


# ---------------------------------------------------------------------------
# Data fetching
# ---------------------------------------------------------------------------

def _fetch_results(target: date) -> pd.DataFrame:
    try:
        resp = requests.get(
            _MLB_SCHEDULE,
            params={"sportId": 1, "date": target.isoformat(),
                    "gameType": "R", "hydrate": "linescore,team"},
            timeout=15,
        )
        resp.raise_for_status()
        dates = resp.json().get("dates", [])
        games = dates[0].get("games", []) if dates else []
    except Exception as exc:
        logger.warning("MLB results fetch failed for %s: %s", target, exc)
        return pd.DataFrame()

    rows = []
    seen: set[int] = set()
    for g in games:
        pk = g.get("gamePk")
        if pk in seen:
            continue
        seen.add(pk)
        if g.get("status", {}).get("abstractGameState") != "Final":
            continue
        ht = _MLB_TO_RETRO.get(g["teams"]["home"]["team"]["abbreviation"],
                               g["teams"]["home"]["team"]["abbreviation"])
        at = _MLB_TO_RETRO.get(g["teams"]["away"]["team"]["abbreviation"],
                               g["teams"]["away"]["team"]["abbreviation"])
        hs = g["teams"]["home"].get("score", np.nan)
        as_ = g["teams"]["away"].get("score", np.nan)
        rows.append({
            "home_team": ht, "visiting_team": at,
            "actual_home_score": hs, "actual_away_score": as_,
            "actual_home_win": int(hs > as_) if pd.notna(hs) and pd.notna(as_) else np.nan,
        })
    return pd.DataFrame(rows)


def _build_odds_lookup(odds_long: pd.DataFrame) -> pd.DataFrame:
    """Reduce historical odds to one best-price row per (home_retro, away_retro)."""
    if odds_long.empty:
        return pd.DataFrame()
    h2h = odds_long[odds_long["market"] == "h2h"].copy()
    h2h["home_retro"] = h2h["home_team"].map(_FULL_TO_RETRO).fillna(h2h["home_team"])
    h2h["away_retro"] = h2h["away_team"].map(_FULL_TO_RETRO).fillna(h2h["away_team"])
    h2h["outcome_retro"] = h2h["outcome_name"].map(_FULL_TO_RETRO).fillna(h2h["outcome_name"])
    h2h["raw_prob"] = h2h["price"].apply(lambda p: american_to_prob(p) if pd.notna(p) else np.nan)

    # no-vig per (event, book)
    grp_tot = h2h.groupby(["event_id", "book"])["raw_prob"].transform("sum")
    h2h["no_vig_prob"] = np.where(grp_tot > 0, h2h["raw_prob"] / grp_tot, np.nan)

    # Keep home-team outcome only, best price across books
    home_side = h2h[h2h["outcome_retro"] == h2h["home_retro"]].copy()
    best = (home_side.sort_values("price", ascending=False)
                     .drop_duplicates(subset=["home_retro", "away_retro"], keep="first")
                     [["home_retro", "away_retro", "book", "price", "no_vig_prob"]]
                     .rename(columns={"book": "best_book", "price": "best_price",
                                      "no_vig_prob": "best_no_vig_prob"}))

    # consensus no-vig (mean across books)
    consensus = (home_side.groupby(["home_retro", "away_retro"])["no_vig_prob"]
                          .mean().rename("consensus_no_vig_prob").reset_index())
    return best.merge(consensus, on=["home_retro", "away_retro"], how="left")


# ---------------------------------------------------------------------------
# Scoring
# ---------------------------------------------------------------------------

def american_to_decimal(american: float) -> float:
    if american >= 100:
        return american / 100 + 1
    return 100 / abs(american) + 1


def _pnl(stake: float, price: float, won: int) -> float:
    if pd.isna(stake) or stake <= 0 or pd.isna(price):
        return 0.0
    dec = american_to_decimal(float(price))
    return stake * (dec - 1) if won == 1 else -stake


_SYNTHETIC_PRICE = -110   # standard US market line used when real odds unavailable


def score_predictions(picks: pd.DataFrame, results: pd.DataFrame,
                      odds_lookup: pd.DataFrame) -> pd.DataFrame:
    merged = picks.copy()

    # Join actual results
    if not results.empty:
        merged = merged.merge(
            results[["home_team", "visiting_team", "actual_home_score",
                      "actual_away_score", "actual_home_win"]],
            on=["home_team", "visiting_team"], how="left",
        )
    else:
        merged["actual_home_score"] = np.nan
        merged["actual_away_score"] = np.nan
        merged["actual_home_win"] = np.nan

    # Join historical odds when available
    if not odds_lookup.empty:
        merged = merged.merge(
            odds_lookup,
            left_on=["home_team", "visiting_team"],
            right_on=["home_retro", "away_retro"],
            how="left",
            suffixes=("_orig", ""),
        )
        for col in ("best_book", "best_price", "best_no_vig_prob", "consensus_no_vig_prob"):
            if f"{col}_orig" in merged.columns:
                merged[col] = merged[col].fillna(merged.get(f"{col}_orig"))
                merged.drop(columns=[f"{col}_orig"], errors="ignore", inplace=True)
        merged.drop(columns=["home_retro", "away_retro"], errors="ignore", inplace=True)

    # Fall back to synthetic -110 line where real odds are missing
    no_price = merged["best_price"].isna() if "best_price" in merged.columns else pd.Series(True, index=merged.index)
    merged["odds_source"] = np.where(no_price, "synthetic_-110", "historical")
    if "best_price" not in merged.columns:
        merged["best_price"] = np.nan
    merged["best_price"] = merged["best_price"].fillna(_SYNTHETIC_PRICE)
    if "best_no_vig_prob" not in merged.columns:
        merged["best_no_vig_prob"] = np.nan
    # no-vig of -110/-110 market is 52.38%
    merged["best_no_vig_prob"] = merged["best_no_vig_prob"].fillna(0.5238)

    merged["edge_vs_best"] = merged["model_win_prob"] - merged["best_no_vig_prob"]

    # Quarter-Kelly stake
    def _kelly(row):
        p = row.get("model_win_prob", np.nan)
        price = row.get("best_price", np.nan)
        if pd.isna(p) or pd.isna(price):
            return 0.0
        dec = american_to_decimal(float(price))
        b = dec - 1
        f = (b * p - (1 - p)) / b
        return max(0.0, f * 0.25)

    merged["kelly_fraction"] = merged.apply(_kelly, axis=1)
    merged["recommended_stake_usd"] = (merged["kelly_fraction"] * config.BANKROLL_USD).round(2)

    merged["pnl_usd"] = merged.apply(
        lambda r: _pnl(r.get("recommended_stake_usd", 0),
                        r.get("best_price", np.nan),
                        r.get("actual_home_win", np.nan))
        if pd.notna(r.get("actual_home_win")) else np.nan,
        axis=1,
    )
    merged["model_correct"] = (
        (merged["model_win_prob"] >= 0.5) == (merged["actual_home_win"] == 1)
    ).where(merged["actual_home_win"].notna())

    return merged


# ---------------------------------------------------------------------------
# Summary tables
# ---------------------------------------------------------------------------

def _summary_at_threshold(df: pd.DataFrame, threshold: float,
                           direction: str = "home") -> dict:
    """Stats for games where model confidence ≥ threshold (home or away side)."""
    if direction == "home":
        mask = df["model_win_prob"] >= threshold
        won_col = "actual_home_win"
    else:
        mask = (1 - df["model_win_prob"]) >= threshold
        won_col = "actual_home_win"

    sub = df[mask].copy()
    n = len(sub)
    scored = sub[won_col].notna().sum()
    correct = sub["model_correct"].sum() if scored > 0 else 0
    staked = sub["recommended_stake_usd"].fillna(0).sum()
    pnl = sub["pnl_usd"].sum() if "pnl_usd" in sub else 0.0
    return {
        "confidence_threshold": f"≥{int(threshold*100)}%",
        "picks": int(n),
        "picks_with_result": int(scored),
        "accuracy": round(correct / scored, 3) if scored > 0 else np.nan,
        "total_staked_usd": round(float(staked), 2),
        "pnl_usd": round(float(pnl), 2) if pd.notna(pnl) else np.nan,
        "roi_pct": round(float(pnl) / float(staked) * 100, 1)
                   if staked > 0 and pd.notna(pnl) else np.nan,
    }


def build_summary(all_games: pd.DataFrame) -> pd.DataFrame:
    all_games = all_games.copy()
    all_games["date_only"] = pd.to_datetime(all_games["date"]).dt.date
    rows = []
    for day, grp in all_games.groupby("date_only"):
        staked = grp["recommended_stake_usd"].fillna(0).sum()
        pnl = grp["pnl_usd"].sum() if "pnl_usd" in grp else 0.0
        scored = int(grp["actual_home_win"].notna().sum())
        correct = grp["model_correct"].sum() if scored > 0 else 0
        rows.append({
            "date": str(day),
            "games": len(grp),
            "games_with_result": scored,
            "model_accuracy": round(correct / scored, 3) if scored > 0 else np.nan,
            "total_staked_usd": round(float(staked), 2),
            "pnl_usd": round(float(pnl), 2) if pd.notna(pnl) else np.nan,
            "roi_pct": round(float(pnl) / float(staked) * 100, 1)
                       if staked > 0 and pd.notna(pnl) else np.nan,
        })
    staked_tot = all_games["recommended_stake_usd"].fillna(0).sum()
    pnl_tot = all_games["pnl_usd"].sum() if "pnl_usd" in all_games else 0.0
    scored_tot = int(all_games["actual_home_win"].notna().sum())
    correct_tot = all_games["model_correct"].sum() if scored_tot > 0 else 0
    rows.append({
        "date": "TOTAL",
        "games": len(all_games),
        "games_with_result": scored_tot,
        "model_accuracy": round(correct_tot / scored_tot, 3) if scored_tot > 0 else np.nan,
        "total_staked_usd": round(float(staked_tot), 2),
        "pnl_usd": round(float(pnl_tot), 2) if pd.notna(pnl_tot) else np.nan,
        "roi_pct": round(float(pnl_tot) / float(staked_tot) * 100, 1)
                   if staked_tot > 0 and pd.notna(pnl_tot) else np.nan,
    })
    return pd.DataFrame(rows)


def build_confidence_table(all_games: pd.DataFrame) -> pd.DataFrame:
    """One row per threshold: picks where model is confident ≥ T in EITHER direction."""
    rows = []
    for t in CONFIDENCE_THRESHOLDS:
        df = all_games.copy()
        # confident = model prob ≥ T OR (1-prob) ≥ T, i.e. abs(prob - 0.5) ≥ T - 0.5
        mask = (df["model_win_prob"] >= t) | ((1 - df["model_win_prob"]) >= t)
        sub = df[mask].copy()
        # For each row decide which "side" the model is backing and whether it won
        sub["model_backed_home"] = sub["model_win_prob"] >= 0.5
        sub["backed_correct"] = (
            (sub["model_backed_home"] & (sub["actual_home_win"] == 1)) |
            (~sub["model_backed_home"] & (sub["actual_home_win"] == 0))
        ).where(sub["actual_home_win"].notna())

        n = len(sub)
        scored = sub["actual_home_win"].notna().sum()
        correct = sub["backed_correct"].sum() if scored > 0 else 0
        staked = sub["recommended_stake_usd"].fillna(0).sum()
        pnl = sub["pnl_usd"].sum() if "pnl_usd" in sub else 0.0
        rows.append({
            "confidence_≥": f"{int(t*100)}%",
            "picks": int(n),
            "picks_scored": int(scored),
            "accuracy": round(correct / scored, 3) if scored > 0 else np.nan,
            "staked_usd": round(float(staked), 2),
            "pnl_usd_(synthetic_-110)": round(float(pnl), 2) if pd.notna(pnl) else np.nan,
            "roi_%_(synthetic_-110)": round(float(pnl) / float(staked) * 100, 1)
                                       if staked > 0 and pd.notna(pnl) else np.nan,
        })
    return pd.DataFrame(rows)


# ---------------------------------------------------------------------------
# Excel styling
# ---------------------------------------------------------------------------

def _style_ws(ws, header_color="1F4E79"):
    hfill = PatternFill("solid", fgColor=header_color)
    hfont = Font(color="FFFFFF", bold=True)
    green = PatternFill("solid", fgColor="C6EFCE")
    red   = PatternFill("solid", fgColor="FFC7CE")

    for cell in ws[1]:
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center")

    col_map = {cell.value: cell.column for cell in ws[1]}
    pnl_col = col_map.get("pnl_usd")
    roi_col = col_map.get("roi_pct")
    correct_col = col_map.get("model_correct")

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column in (pnl_col, roi_col) and isinstance(cell.value, (int, float)):
                cell.fill = green if cell.value > 0 else (red if cell.value < 0 else PatternFill())
            if correct_col and cell.column == correct_col:
                cell.fill = (green if cell.value is True
                             else (red if cell.value is False else PatternFill()))

    for col_cells in ws.columns:
        width = max((len(str(c.value or "")) for c in col_cells), default=8)
        ws.column_dimensions[col_cells[0].column_letter].width = min(width + 2, 32)


def _add_accuracy_chart(ws, n_data_rows: int):
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    if "model_accuracy" not in headers or "date" not in headers:
        return
    acc_col  = headers.index("model_accuracy") + 1
    date_col = headers.index("date") + 1
    chart = BarChart()
    chart.type = "col"
    chart.title = "Model Accuracy by Day"
    chart.y_axis.title = "Accuracy"
    chart.y_axis.numFmt = "0%"
    chart.width = 18
    chart.height = 10
    data = Reference(ws, min_col=acc_col, min_row=1, max_row=n_data_rows + 1)
    cats = Reference(ws, min_col=date_col, min_row=2, max_row=n_data_rows + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.series[0].graphicalProperties.solidFill = "1F4E79"
    ws.add_chart(chart, "I2")


def _add_confidence_chart(ws, n_rows: int):
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    for col_name, anchor, color in [
        ("accuracy",     "I2",  "1F4E79"),
        ("roi_pct",      "I18", "375623"),
    ]:
        if col_name not in headers:
            continue
        col_idx = headers.index(col_name) + 1
        thresh_col = headers.index("confidence_≥") + 1
        chart = BarChart()
        chart.type = "col"
        chart.title = f"{'Accuracy' if col_name=='accuracy' else 'ROI %'} by Confidence Threshold"
        chart.y_axis.numFmt = "0%" if col_name == "accuracy" else "0.0"
        chart.width = 16
        chart.height = 10
        data = Reference(ws, min_col=col_idx, min_row=1, max_row=n_rows + 1)
        cats = Reference(ws, min_col=thresh_col, min_row=2, max_row=n_rows + 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.series[0].graphicalProperties.solidFill = color
        ws.add_chart(chart, anchor)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def run(start: date, end: date) -> Path:
    # 1. Collect picks files
    picks_dir = config.PICKS_DIR
    all_frames: list[pd.DataFrame] = []
    dates: list[date] = []
    d = start
    while d <= end:
        p = picks_dir / f"picks_{d.isoformat()}.csv"
        if p.exists():
            df = pd.read_csv(p)
            if not df.empty:
                df["_date"] = d
                all_frames.append(df)
                dates.append(d)
        d += timedelta(days=1)

    if not all_frames:
        logger.error("no picks files found for %s to %s", start, end)
        return Path()

    # 2. Try to fetch historical odds (requires paid Odds API plan)
    print(f"Attempting to fetch historical odds for {len(dates)} dates…")
    hist_odds = fetch_historical_dates(dates)
    if hist_odds.empty:
        print("  → Historical odds unavailable on current API plan.")
        print("  → Using synthetic -110 / -110 lines as conservative ROI baseline.")
        print("  → Upgrade to the-odds-api.com Hobby plan ($29/mo) for real historical lines.\n")
    else:
        logger.info("fetched %d historical odds rows across %d dates", len(hist_odds), len(dates))

    # 3. Score each day
    scored_frames: list[pd.DataFrame] = []
    for d, picks in zip(dates, all_frames):
        results    = _fetch_results(d)
        # Filter odds to this date's games
        if not hist_odds.empty and "commence_time" in hist_odds.columns:
            day_odds = hist_odds[
                pd.to_datetime(hist_odds["commence_time"]).dt.date == d
            ].copy()
        else:
            day_odds = hist_odds.copy()
        odds_lookup = _build_odds_lookup(day_odds)
        scored = score_predictions(picks.drop(columns=["_date"]), results, odds_lookup)
        scored["date"] = d.isoformat()
        scored_frames.append(scored)
        staked = scored["recommended_stake_usd"].fillna(0).sum()
        pnl    = scored["pnl_usd"].sum()
        logger.info("%s: %d games, %d results, %d odds rows, staked=%.2f, P&L=%.2f",
                    d, len(picks), len(results), len(day_odds), staked, pnl)

    combined = pd.concat(scored_frames, ignore_index=True)
    summary    = build_summary(combined)
    conf_table = build_confidence_table(combined)

    # 4. Print summaries
    print("\n=== Daily Performance ===")
    print(summary.to_string(index=False))
    print("\n=== Accuracy & ROI by Confidence Threshold (home-team picks) ===")
    print(conf_table.to_string(index=False))

    # 5. Write Excel
    label  = f"{start.isoformat()}_to_{end.isoformat()}"
    csv_out = picks_dir / f"backtest_{label}.csv"
    xl_out  = picks_dir / f"backtest_{label}.xlsx"
    combined.to_csv(csv_out, index=False)

    with pd.ExcelWriter(xl_out, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="Daily_Summary", index=False)
        _style_ws(writer.sheets["Daily_Summary"], header_color="375623")
        n_days = len(summary) - 1   # exclude TOTAL row
        _add_accuracy_chart(writer.sheets["Daily_Summary"], n_days)

        conf_table.to_excel(writer, sheet_name="Confidence_Filter", index=False)
        _style_ws(writer.sheets["Confidence_Filter"], header_color="7B2C2C")
        _add_confidence_chart(writer.sheets["Confidence_Filter"], len(conf_table))

        combined.to_excel(writer, sheet_name="All_Games", index=False)
        _style_ws(writer.sheets["All_Games"])

        combined["_date"] = pd.to_datetime(combined["date"]).dt.date
        for day, grp in combined.groupby("_date"):
            sheet = pd.Timestamp(day).strftime("%b_%d")
            grp.drop(columns=["_date"]).to_excel(writer, sheet_name=sheet, index=False)
            _style_ws(writer.sheets[sheet])
        combined.drop(columns=["_date"], inplace=True)

    logger.info("wrote backtest to %s", xl_out)
    return xl_out


def main():
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
    parser = argparse.ArgumentParser()
    parser.add_argument("--start", default="2026-05-08")
    parser.add_argument("--end",   default="2026-05-14")
    args = parser.parse_args()
    run(date.fromisoformat(args.start), date.fromisoformat(args.end))


if __name__ == "__main__":
    main()
